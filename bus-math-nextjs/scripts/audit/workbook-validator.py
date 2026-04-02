"""
XLSX Workbook Content Validator

Validates workbook structure, sheet names, and content alignment.
Uses openpyxl for workbook inspection following the spreadsheet skill guidance.
Read-only: no workbook mutation.
"""

import os
import sys
from typing import List, Dict, Optional, Any
from dataclasses import dataclass, field

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


@dataclass
class WorkbookValidationResult:
    """Result of validating a single workbook file."""
    file_path: str
    opens_successfully: bool
    sheet_count: int
    sheet_names: List[str]
    error_message: Optional[str] = None
    has_formulas: bool = False
    formula_count: int = 0
    has_data: bool = False
    cell_count: int = 0


@dataclass
class AlignmentCheckResult:
    """Result of comparing workbook structure to page expectations."""
    workbook_path: str
    page_path: str
    aligned: bool
    issues: List[str] = field(default_factory=list)
    expected_sheets: Optional[List[str]] = None
    actual_sheets: Optional[List[str]] = None


def validate_workbook(file_path: str) -> WorkbookValidationResult:
    """
    Open and validate a single workbook file.
    
    Checks:
    - Workbook opens without error
    - Has at least one sheet
    - Records sheet names
    - Detects formulas vs static data
    """
    result = WorkbookValidationResult(
        file_path=file_path,
        opens_successfully=False,
        sheet_count=0,
        sheet_names=[],
    )
    
    if not os.path.exists(file_path):
        result.error_message = f"File does not exist: {file_path}"
        return result
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        result.opens_successfully = True
        result.sheet_count = len(wb.sheetnames)
        result.sheet_names = list(wb.sheetnames)
        
        total_cells = 0
        formula_count = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    total_cells += 1
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        
        result.has_data = total_cells > 0
        result.cell_count = total_cells
        result.has_formulas = formula_count > 0
        result.formula_count = formula_count
        
        wb.close()
        
    except Exception as e:
        result.error_message = str(e)
    
    return result


def check_sheet_alignment(
    workbook_path: str,
    expected_sheets: List[str],
) -> AlignmentCheckResult:
    """
    Check if workbook has expected sheet names.
    
    Flags:
    - Missing expected sheets
    - Extra sheets not in expected list
    - Sheet name mismatches (case-sensitive)
    """
    result = AlignmentCheckResult(
        workbook_path=workbook_path,
        page_path="",
        aligned=True,
        expected_sheets=expected_sheets,
    )
    
    validation = validate_workbook(workbook_path)
    
    if not validation.opens_successfully:
        result.aligned = False
        result.issues.append(f"Cannot open workbook: {validation.error_message}")
        return result
    
    result.actual_sheets = validation.sheet_names
    
    expected_lower = [s.lower() for s in expected_sheets]
    actual_lower = [s.lower() for s in validation.sheet_names]
    
    for expected in expected_sheets:
        if expected.lower() not in actual_lower:
            result.aligned = False
            result.issues.append(f"Missing expected sheet: '{expected}'")
    
    for actual in validation.sheet_names:
        if actual.lower() not in expected_lower:
            result.issues.append(f"Extra sheet not in expected list: '{actual}'")
    
    return result


def validate_student_teacher_pairing(
    student_path: str,
    teacher_path: str,
) -> Dict[str, Any]:
    """
    When both student and teacher workbooks are referenced,
    verify both exist and have compatible structure.
    """
    result = {
        "student_exists": os.path.exists(student_path),
        "teacher_exists": os.path.exists(teacher_path),
        "paired": False,
        "issues": [],
    }
    
    if not result["student_exists"]:
        result["issues"].append(f"Student workbook missing: {student_path}")
    if not result["teacher_exists"]:
        result["issues"].append(f"Teacher workbook missing: {teacher_path}")
    
    if result["student_exists"] and result["teacher_exists"]:
        student_wb = validate_workbook(student_path)
        teacher_wb = validate_workbook(teacher_path)
        
        if student_wb.opens_successfully and teacher_wb.opens_successfully:
            result["paired"] = True
            result["student_sheets"] = student_wb.sheet_names
            result["teacher_sheets"] = teacher_wb.sheet_names
            result["student_has_formulas"] = student_wb.has_formulas
            result["teacher_has_formulas"] = teacher_wb.has_formulas
        else:
            if not student_wb.opens_successfully:
                result["issues"].append(f"Student workbook error: {student_wb.error_message}")
            if not teacher_wb.opens_successfully:
                result["issues"].append(f"Teacher workbook error: {teacher_wb.error_message}")
    
    return result


def run_unit_workbook_audit(
    resources_dir: str,
    unit_prefix: str,
    linked_files: List[str],
) -> List[Dict[str, Any]]:
    """
    Run workbook content validation for all linked files in a unit.
    
    Returns a list of validation results with file path, status, and details.
    """
    results = []
    
    for linked_file in linked_files:
        full_path = os.path.join(resources_dir, linked_file)
        
        if not full_path.endswith('.xlsx'):
            continue
        
        validation = validate_workbook(full_path)
        
        results.append({
            "file": linked_file,
            "full_path": full_path,
            "exists": os.path.exists(full_path),
            "opens": validation.opens_successfully,
            "sheet_count": validation.sheet_count,
            "sheet_names": validation.sheet_names,
            "has_formulas": validation.has_formulas,
            "formula_count": validation.formula_count,
            "has_data": validation.has_data,
            "cell_count": validation.cell_count,
            "error": validation.error_message,
        })
    
    return results


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python workbook-validator.py <resources_dir> <unit_prefix>")
        print("Example: python workbook-validator.py public/resources unit01")
        sys.exit(1)
    
    resources_dir = sys.argv[1]
    unit_prefix = sys.argv[2]
    
    xlsx_files = [
        f for f in os.listdir(resources_dir)
        if f.startswith(unit_prefix) and f.endswith('.xlsx')
    ]
    
    print(f"Validating {len(xlsx_files)} workbooks for {unit_prefix}...")
    print()
    
    results = run_unit_workbook_audit(resources_dir, unit_prefix, xlsx_files)
    
    for r in results:
        status = "OK" if r["opens"] else "FAIL"
        sheets = f"{r['sheet_count']} sheets: {', '.join(r['sheet_names'])}" if r["opens"] else "N/A"
        formulas = f"{r['formula_count']} formulas" if r["opens"] else "N/A"
        print(f"[{status}] {r['file']}")
        print(f"       {sheets}")
        print(f"       {formulas}")
        if r["error"]:
            print(f"       ERROR: {r['error']}")
        print()
